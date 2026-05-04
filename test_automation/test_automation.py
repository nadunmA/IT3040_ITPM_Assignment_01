from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from pathlib import Path
import sys
import openpyxl
from openpyxl.cell.cell import MergedCell

ROOT_DIR = Path(__file__).resolve().parent.parent
TESTS_DIR = ROOT_DIR / "test_automation"

DEFAULT_EXCEL_CANDIDATES = [
    str(TESTS_DIR / "IT23192850.xlsx"),
]

DEFAULT_SHEET_NAME = " Test cases"
DEFAULT_FRONTEND_URL = os.getenv("FRONTEND_URL", "https://tmrtools.com/")

DEFAULT_INPUT_COLUMN_CANDIDATES = [
    "Singlish", "Input", "Singlish Input", "Test Input",
    "Source", "Sentence", "Text",
]

DEFAULT_EXPECTED_COLUMN_CANDIDATES = [
    "Sinhala", "Expected_Output", "Expected Output",
    "Expected output", "Expected", "Expected Sinhala",
]

DEFAULT_ACTUAL_COLUMN_CANDIDATES = [
    "Actual_Output", "Actual Output", "Actual output", "Actual",
]

DEFAULT_STATUS_COLUMN_CANDIDATES = [
    "Status", "Result", "Pass/Fail", "Pass Fail",
]

DEFAULT_WAIT_MS = 15000
DEFAULT_RETRIES = 10
DEFAULT_RETRY_WAIT_MS = 2000
DEFAULT_TYPE_DELAY_MS = 50
DEFAULT_TIMEOUT_MS = 60000
DEFAULT_SLOW_MO_MS = 300


def _configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass


def _pick_existing_path(candidates):
    for p in candidates:
        if p and os.path.exists(p):
            return p
    return candidates[0] if candidates else None


def _resolve_path(p: str | None) -> str | None:
    if not p:
        return None
    path = Path(p)
    if path.is_absolute():
        return str(path)
    root_candidate = (ROOT_DIR / path).resolve()
    if root_candidate.exists():
        return str(root_candidate)
    tests_candidate = (TESTS_DIR / path).resolve()
    if tests_candidate.exists():
        return str(tests_candidate)
    return str(root_candidate)


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _header_values(ws, row_index: int) -> list:
    max_col = max(1, int(ws.max_column or 1))
    return [ws.cell(row=row_index, column=c).value for c in range(1, max_col + 1)]


def _find_header_row(ws, max_scan_rows: int) -> int:
    expected_tokens = {_normalize_header(v) for v in DEFAULT_EXPECTED_COLUMN_CANDIDATES}
    best_score = -1
    best_row = 1

    scan_limit = max(1, min(int(max_scan_rows), int(ws.max_row or 1)))

    for r in range(1, scan_limit + 1):
        values = _header_values(ws, r)
        texts = [v for v in values if isinstance(v, str) and v.strip()]
        norms = {_normalize_header(v) for v in texts}

        if "input" in norms and norms & expected_tokens:
            score = len(norms & expected_tokens) + (3 if "input" in norms else 0)
            if score > best_score:
                best_score = score
                best_row = r

    return best_row


def _merged_top_left_cell(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell

    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)

    return ws.cell(row=row, column=col)


def _is_top_left_of_merged_cell(ws, row: int, col: int) -> bool:
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return True

    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row == row and rng.min_col == col

    return True


def _set_cell_value(ws, row: int, col: int, value):
    cell = _merged_top_left_cell(ws, row, col)
    cell.value = value


def _find_column_index(header_values: list, requested_name: str | None, candidates: list[str]) -> int | None:
    indexed = [(i, str(v)) for i, v in enumerate(header_values, start=1) if v is not None]

    norm_to_index = {}
    for i, v in indexed:
        n = _normalize_header(v)
        if n and n not in norm_to_index:
            norm_to_index[n] = i

    def match(name: str) -> int | None:
        n = _normalize_header(name)
        if not n:
            return None

        if n in norm_to_index:
            return norm_to_index[n]

        for i, v in indexed:
            hv = _normalize_header(v)
            if n in hv or hv in n:
                return i

        return None

    if requested_name:
        found = match(requested_name)
        if found:
            return found

    for c in candidates:
        found = match(c)
        if found:
            return found

    return None


def _last_header_col(header_values: list) -> int:
    last = 0
    for i, v in enumerate(header_values, start=1):
        if v is not None and str(v).strip():
            last = i
    return last


def _ensure_column(ws, header_row: int, header_values: list, desired_name: str) -> int:
    found = _find_column_index(header_values, desired_name, [])
    if found:
        return found

    col = _last_header_col(header_values) + 1
    ws.cell(row=header_row, column=col).value = desired_name

    while len(header_values) < col - 1:
        header_values.append(None)
    header_values.append(desired_name)

    return col


def _dismiss_overlays(page):
    patterns = [
        r"Accept", r"Accept all", r"I Agree", r"Agree", r"OK", r"Got it",
        r"Close", r"×"
    ]

    for pattern in patterns:
        try:
            btn = page.get_by_role("button", name=re.compile(pattern, re.IGNORECASE)).first
            if btn.is_visible():
                btn.click(timeout=2000)
                page.wait_for_timeout(300)
        except Exception:
            pass


def _clear_textarea(page, locator):
    try:
        locator.click(timeout=3000)
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")
    except Exception:
        pass

    try:
        locator.fill("")
    except Exception:
        pass

    try:
        locator.evaluate("""
            el => {
                if ('value' in el) el.value = '';
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
            }
        """)
    except Exception:
        pass


def _ensure_input_value(page, input_locator, text: str, type_delay_ms: int):
    _clear_textarea(page, input_locator)

    if type_delay_ms > 0:
        input_locator.click(timeout=3000)
        input_locator.type(text, delay=type_delay_ms)
    else:
        input_locator.fill(text)

    page.wait_for_timeout(300)

    try:
        current = input_locator.input_value()
        if current.strip() != text.strip():
            _clear_textarea(page, input_locator)
            input_locator.fill(text)
    except Exception:
        pass


def _read_output(output_locator) -> str:
    methods = [
        lambda: output_locator.input_value(),
        lambda: output_locator.inner_text(),
        lambda: output_locator.text_content(),
        lambda: output_locator.evaluate("el => 'value' in el ? el.value : el.innerText")
    ]

    for method in methods:
        try:
            value = method()
            if value:
                value = str(value).strip()
                if value:
                    return value
        except Exception:
            pass

    return ""


def _find_tmrtools_locators(page, timeout_ms: int):
    deadline = time.time() + timeout_ms / 1000

    while time.time() < deadline:
        _dismiss_overlays(page)

        try:
            input_locator = page.locator(
                'textarea[placeholder*="Singlish"], '
                'textarea[placeholder*="English"], '
                'textarea'
            ).first

            if input_locator.count() > 0 and input_locator.is_visible():
                pass
            else:
                page.wait_for_timeout(500)
                continue

            button = page.get_by_role(
                "button",
                name=re.compile(r"(Translate|Convert|Transliterate|Generate)", re.IGNORECASE)
            ).first

            output_candidates = [
                page.locator("textarea").nth(1),
                page.locator('[contenteditable="true"]').first,
                page.locator("div").filter(has_text=re.compile(r"[\u0D80-\u0DFF]+")).first,
                page.locator("p").filter(has_text=re.compile(r"[\u0D80-\u0DFF]+")).first,
            ]

            output_locator = None
            for loc in output_candidates:
                try:
                    if loc.count() > 0:
                        output_locator = loc
                        break
                except Exception:
                    pass

            if output_locator is None:
                output_locator = page.locator("body")

            return input_locator, output_locator, button

        except Exception:
            page.wait_for_timeout(500)

    raise RuntimeError("Could not find tmrtools.com input/output elements.")


def _parse_args():
    parser = argparse.ArgumentParser()

    parser.add_argument("--excel", default=_pick_existing_path(DEFAULT_EXCEL_CANDIDATES))
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--header-row", type=int, default=0)
    parser.add_argument("--max-header-scan-rows", type=int, default=30)

    parser.add_argument("--input-col", default=None)
    parser.add_argument("--expected-col", default=None)
    parser.add_argument("--actual-col", default=None)
    parser.add_argument("--status-col", default=None)

    parser.add_argument("--url", default=DEFAULT_FRONTEND_URL)
    parser.add_argument("--output", default=None)

    parser.add_argument("--save-every", type=int, default=1)
    parser.add_argument("--headless", action="store_true", default=False)

    parser.add_argument("--wait-ms", type=int, default=DEFAULT_WAIT_MS)
    parser.add_argument("--retries", type=int, default=DEFAULT_RETRIES)
    parser.add_argument("--retry-wait-ms", type=int, default=DEFAULT_RETRY_WAIT_MS)
    parser.add_argument("--type-delay-ms", type=int, default=DEFAULT_TYPE_DELAY_MS)
    parser.add_argument("--timeout-ms", type=int, default=DEFAULT_TIMEOUT_MS)
    parser.add_argument("--slow-mo-ms", type=int, default=DEFAULT_SLOW_MO_MS)
    parser.add_argument("--keep-open", action="store_true", default=False)

    return parser.parse_args()


def run_test():
    _configure_stdout()
    args = _parse_args()

    args.excel = _resolve_path(args.excel)
    args.output = _resolve_path(args.output) if args.output else args.excel

    if not args.excel or not os.path.exists(args.excel):
        print(f"Error: File '{args.excel}' not found.")
        return

    try:
        wb = openpyxl.load_workbook(args.excel)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    if args.sheet and args.sheet in wb.sheetnames:
        ws = wb[args.sheet]
    else:
        ws = wb.active

    header_row = int(args.header_row or 0)
    if header_row <= 0:
        header_row = _find_header_row(ws, args.max_header_scan_rows)

    header_values = _header_values(ws, header_row)

    input_col_idx = _find_column_index(header_values, args.input_col, DEFAULT_INPUT_COLUMN_CANDIDATES)
    expected_col_idx = _find_column_index(header_values, args.expected_col, DEFAULT_EXPECTED_COLUMN_CANDIDATES)

    if not input_col_idx:
        print("Error: Could not resolve input column.")
        print(f"Header row: {header_row}")
        print(f"Available columns: {header_values}")
        return

    actual_col_idx = _find_column_index(header_values, args.actual_col, DEFAULT_ACTUAL_COLUMN_CANDIDATES)
    status_col_idx = _find_column_index(header_values, args.status_col, DEFAULT_STATUS_COLUMN_CANDIDATES)

    actual_col_idx = actual_col_idx or _ensure_column(ws, header_row, header_values, args.actual_col or "Actual output")
    status_col_idx = status_col_idx or _ensure_column(ws, header_row, header_values, args.status_col or "Status")

    rows_total = max(0, int(ws.max_row or 0) - header_row)
    print(f"Starting Frontend-Only test with {rows_total} rows...")
    print(f"URL: {args.url}")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=args.headless,
            slow_mo=max(0, int(args.slow_mo_ms))
        )

        page = browser.new_page()
        page.set_default_timeout(max(1000, int(args.timeout_ms)))

        try:
            page.goto(args.url, wait_until="domcontentloaded", timeout=args.timeout_ms)

            try:
                page.wait_for_load_state("networkidle", timeout=args.timeout_ms)
            except Exception:
                pass

            page.wait_for_selector("textarea", timeout=args.timeout_ms)
            print("Frontend loaded successfully.")

        except Exception as e:
            print(f"Error loading frontend: {e}")
            browser.close()
            return

        try:
            input_locator, output_locator, action_locator = _find_tmrtools_locators(page, args.timeout_ms)
            print("UI elements found successfully.")
        except Exception as e:
            print(f"Error locating UI elements: {e}")
            browser.close()
            return

        processed = 0

        for row_index in range(header_row + 1, int(ws.max_row or 0) + 1):
            if not _is_top_left_of_merged_cell(ws, row_index, input_col_idx):
                continue

            input_value = _merged_top_left_cell(ws, row_index, input_col_idx).value
            singlish_input = str(input_value).strip() if input_value is not None else ""

            if not singlish_input:
                continue

            expected_value = (
                _merged_top_left_cell(ws, row_index, expected_col_idx).value
                if expected_col_idx else None
            )
            expected_sinhala = str(expected_value).strip() if expected_value is not None else ""

            print(f"Testing [Row {row_index}]: {singlish_input}")

            try:
                _dismiss_overlays(page)

                prev_output = _read_output(output_locator)

                _ensure_input_value(
                    page,
                    input_locator,
                    singlish_input,
                    int(args.type_delay_ms)
                )

                try:
                    if action_locator and action_locator.count() > 0 and action_locator.is_visible():
                        action_locator.click(timeout=5000)
                except Exception:
                    pass

                page.wait_for_timeout(max(0, int(args.wait_ms)))

                actual_output = ""

                for _ in range(max(1, int(args.retries))):
                    current = _read_output(output_locator)

                    if current and current != prev_output:
                        actual_output = current
                        break

                    if current and not prev_output:
                        actual_output = current
                        break

                    page.wait_for_timeout(max(0, int(args.retry_wait_ms)))

                _set_cell_value(ws, row_index, actual_col_idx, actual_output)

                if expected_sinhala:
                    status = "PASS" if actual_output == expected_sinhala else "FAIL"
                else:
                    status = "COLLECTED"

                _set_cell_value(ws, row_index, status_col_idx, status)

                print(f"  -> {status}")
                print(f"  Actual: {actual_output}")

                processed += 1

                if args.save_every and processed % int(args.save_every) == 0:
                    wb.save(args.output)

            except Exception as e:
                print(f"Error in UI interaction row {row_index}: {e}")

                try:
                    _set_cell_value(ws, row_index, status_col_idx, "UI Error")
                except Exception:
                    pass

                if args.save_every:
                    try:
                        wb.save(args.output)
                    except Exception:
                        pass

        if args.keep_open and not args.headless:
            try:
                wb.save(args.output)
            except Exception:
                pass

            print("Keeping browser open. Press CTRL+C to stop.")

            try:
                while True:
                    page.wait_for_timeout(1000)
            except KeyboardInterrupt:
                pass

        browser.close()

    try:
        wb.save(args.output)
    except Exception as e:
        print(f"Error saving output file '{args.output}': {e}")
        return

    print(f"Test completed. Results saved to {args.output}")


if __name__ == "__main__":
    run_test()